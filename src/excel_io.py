"""
Excel 読み書きモジュール

元Excelをコピーし、Sheet1のJANコードとSheet2のKeepaデータを
マッチングしてSheet3に数式ごと自動生成する。
"""

from __future__ import annotations

import io
import shutil
from pathlib import Path
import openpyxl
import pandas as pd


def read_wholesaler_excel(file_buffer) -> pd.DataFrame:
    """
    卸問屋Excelから JAN コード・品名・卸価格などを読み取る。
    ヘッダー名で列を自動検出するため、列位置が異なるExcelにも対応。
    """
    raw = pd.read_excel(
        file_buffer,
        sheet_name=0,
        header=0,
        dtype=str,
        engine="openpyxl",
    )

    # ヘッダー名から列を自動検出
    col_map = {}
    for col in raw.columns:
        col_upper = str(col).strip().upper()
        if "JAN" in col_upper and "jan_code" not in col_map:
            col_map["jan_code"] = col
        elif col_upper in ("品名", "商品名") and "product_name_jp" not in col_map:
            col_map["product_name_jp"] = col
        elif col_upper in ("品番", "型番") and "part_number" not in col_map:
            col_map["part_number"] = col
        elif col_upper in ("定価", "上代") and "retail_price" not in col_map:
            col_map["retail_price"] = col
        elif col_upper in ("卸", "卸価格", "仕入値", "仕入れ値") and "wholesale_price" not in col_map:
            col_map["wholesale_price"] = col
        elif col_upper in ("箱入数", "出荷単位", "入数", "ロット") and "box_qty" not in col_map:
            col_map["box_qty"] = col

    # JAN列が見つからない場合、13桁数字が多い列を探す
    if "jan_code" not in col_map:
        for col in raw.columns:
            vals = raw[col].fillna("").str.strip()
            digit_match = vals.str.match(r"^\d{8,13}$")
            if digit_match.sum() > len(raw) * 0.3:
                col_map["jan_code"] = col
                break

    if "jan_code" not in col_map:
        return pd.DataFrame()

    result = pd.DataFrame()
    result["jan_code"] = raw[col_map["jan_code"]].fillna("").str.strip()
    result["product_name_jp"] = raw[col_map.get("product_name_jp", result.columns[0])].fillna("") if "product_name_jp" in col_map else ""
    result["part_number"] = raw[col_map.get("part_number", result.columns[0])].fillna("") if "part_number" in col_map else ""
    result["retail_price"] = pd.to_numeric(raw[col_map["retail_price"]], errors="coerce").fillna(0) if "retail_price" in col_map else 0
    result["wholesale_price"] = pd.to_numeric(raw[col_map["wholesale_price"]], errors="coerce").fillna(0) if "wholesale_price" in col_map else 0
    result["box_qty"] = pd.to_numeric(raw[col_map["box_qty"]], errors="coerce").fillna(1).astype(int) if "box_qty" in col_map else 1

    result = result[result["jan_code"].str.match(r"^\d{8,}$", na=False)].copy()
    return result.reset_index(drop=True)


def read_keepa_export_as_results(file_buffer, sheet_name: str | int = 0) -> dict[str, dict]:
    """
    Keepaエクスポートファイルを読み込み、EAN→商品データの辞書を返す。
    列をヘッダー名で自動検出するため、異なるKeepaエクスポート形式に対応。
    """
    raw = pd.read_excel(
        file_buffer, sheet_name=sheet_name,
        header=0, dtype=str, engine="openpyxl",
    )

    # ヘッダー名で列を自動検出
    col_idx = {}
    for i, col in enumerate(raw.columns):
        c = str(col).strip()
        cu = c.upper()
        if "EAN" in cu and "ean" not in col_idx:
            col_idx["ean"] = i
        if "ASIN" == cu and "asin" not in col_idx:
            col_idx["asin"] = i
        if "商品名" in c and "title" not in col_idx and "ランキング" not in c:
            col_idx["title"] = i
        if "Buy Box" in c and "現在価格" in c and "buybox" not in col_idx and "中古" not in c:
            col_idx["buybox"] = i
        if "Buy Box" in c and "セラー" in c and "bb_seller" not in col_idx:
            col_idx["bb_seller"] = i
        if "FBA手数料" == c.strip() and "total_fee_col" not in col_idx:
            col_idx["total_fee_col"] = i
        if "FBA Pick" in c and "fba_fee" not in col_idx:
            col_idx["fba_fee"] = i
        if "紹介料" in c and "現在" not in c and "referral" not in col_idx:
            col_idx["referral"] = i
        if "重さ" in c and "パッケージ" in c and "weight" not in col_idx:
            col_idx["weight"] = i
        if "30日間の減少" in c and "drops30" not in col_idx:
            col_idx["drops30"] = i
        if "90日間の減少" in c and "drops90" not in col_idx:
            col_idx["drops90"] = i
        if "合計オファー数" in c and "total_offers" not in col_idx:
            col_idx["total_offers"] = i
        if "新品アイテム数 FBA" in c and "現在" in c and "fba_count" not in col_idx:
            col_idx["fba_count"] = i
        if "新品アイテム数 FBM" in c and "現在" in c and "fbm_count" not in col_idx:
            col_idx["fbm_count"] = i
        if "Imported by Code" in c and "ean" not in col_idx:
            col_idx["ean"] = i
        if "Amazon" in c and "現在価格" in c and "Buy Box" not in c and "amazon_price" not in col_idx:
            col_idx["amazon_price"] = i
        if "% Amazon 90" in c and "amazon_bb_pct" not in col_idx:
            col_idx["amazon_bb_pct"] = i

    # EAN列フォールバック
    if "ean" not in col_idx:
        for i in range(min(len(raw.columns), 100)):
            vals = raw.iloc[:, i].fillna("").str.strip()
            if vals.str.match(r"^\d{8,13}$").sum() > len(raw) * 0.3:
                col_idx["ean"] = i
                break

    if "ean" not in col_idx:
        return {}

    def sf(val):
        try:
            v = float(val)
            return v if v > 0 else None
        except (TypeError, ValueError):
            return None

    def si(val):
        try:
            return int(float(val))
        except (TypeError, ValueError):
            return None

    def get(row, key):
        idx = col_idx.get(key)
        return row.iloc[idx] if idx is not None and idx < len(row) else None

    results: dict[str, dict] = {}
    best_price: dict[str, float] = {}

    for _, row in raw.iterrows():
        ean_raw = str(get(row, "ean") or "").strip()
        if not ean_raw:
            continue

        buy_box = sf(get(row, "buybox"))
        # Amazon手数料: FBA手数料列（合算値）を優先、なければFBA Pick&Pack+紹介料
        total_fee = sf(get(row, "total_fee_col"))
        if total_fee is None:
            fba_fee = sf(get(row, "fba_fee")) or 0
            referral = sf(get(row, "referral")) or 0
            total_fee = (fba_fee + referral) if (fba_fee + referral) > 0 else None

        amazon_price = sf(get(row, "amazon_price"))
        amazon_bb_pct = sf(get(row, "amazon_bb_pct"))
        is_amazon_selling = (amazon_price is not None) or (amazon_bb_pct is not None and amazon_bb_pct > 0)

        product = {
            "found": buy_box is not None and buy_box > 0,
            "asin": str(get(row, "asin") or "").strip(),
            "title": str(get(row, "title") or "").strip(),
            "buy_box_price_usd": buy_box,
            "buy_box_seller": str(get(row, "bb_seller") or "").strip(),
            "fba_seller_count": si(get(row, "total_offers")) or si(get(row, "fba_count")) or 0,
            "fbm_seller_count": si(get(row, "fbm_count")) or 0,
            "sales_rank_drops_30": si(get(row, "drops30")),
            "package_weight_g": si(get(row, "weight")),
            "total_amazon_fee_usd": total_fee,
            "amazon_price_usd": amazon_price,
            "amazon_bb_pct_90": amazon_bb_pct,
            "is_amazon_selling": is_amazon_selling,
        }

        for ean in ean_raw.split(","):
            ean = ean.strip().replace(" ", "")
            if not ean or not ean.isdigit() or len(ean) < 8:
                continue
            old_price = best_price.get(ean, -1)
            if (buy_box or 0) > old_price:
                results[ean] = product
                best_price[ean] = buy_box or 0

    return results


def build_ean_to_sheet2_row(file_buffer, sheet_name: str | int = 1) -> dict[str, int]:
    """
    Sheet2（Keepaエクスポート）を読み、EAN → Sheet2の行番号（1-indexed）のマッピングを作る。
    EAN列とBuyBox列をヘッダー名で自動検出。
    """
    raw = pd.read_excel(
        file_buffer, sheet_name=sheet_name,
        header=0, dtype=str, engine="openpyxl",
    )

    # EAN列を自動検出（ヘッダー名 → 13桁数字フォールバック）
    ean_col_idx = None
    buybox_col_idx = None
    for i, col in enumerate(raw.columns):
        col_str = str(col).strip()
        if ean_col_idx is None and ("Imported by Code" in col_str or
                ("EAN" in col_str.upper() and "商品コード" in col_str)):
            ean_col_idx = i
        if buybox_col_idx is None and "Buy Box" in col_str and "現在価格" in col_str and "中古" not in col_str:
            buybox_col_idx = i

    # EAN列が見つからなければ13桁数字が多い列を探す
    if ean_col_idx is None:
        for i in range(min(len(raw.columns), 100)):
            vals = raw.iloc[:, i].fillna("").str.strip()
            if vals.str.match(r"^\d{8,13}$").sum() > len(raw) * 0.3:
                ean_col_idx = i
                break

    if ean_col_idx is None:
        return {}

    # BuyBox列が見つからなければindex 14を使う（元フォーマット互換）
    if buybox_col_idx is None:
        buybox_col_idx = 14 if len(raw.columns) > 14 else None

    ean_to_row: dict[str, int] = {}
    ean_to_price: dict[str, float] = {}

    for idx, row in raw.iterrows():
        ean_raw = str(row.iloc[ean_col_idx]).strip() if pd.notna(row.iloc[ean_col_idx]) else ""
        if not ean_raw:
            continue
        excel_row = idx + 2
        buy_box = _safe_float(row.iloc[buybox_col_idx]) if buybox_col_idx is not None else 0
        buy_box = buy_box or 0

        for ean in ean_raw.split(","):
            ean = ean.strip().replace(" ", "")
            if not ean or not ean.isdigit() or len(ean) < 8:
                continue
            old_price = ean_to_price.get(ean, -1)
            if buy_box > old_price:
                ean_to_row[ean] = excel_row
                ean_to_price[ean] = buy_box

    return ean_to_row


def generate_sheet3(
    source_path: str | Path,
    output_path: str | Path,
    sheet2_name: str = "2026-01-13",
):
    """
    元Excelをコピーし、Sheet1のJANコードをSheet2とマッチングして
    Sheet3にSheet2への参照数式を自動生成する。

    Args:
        source_path: 元Excelファイルのパス
        output_path: 出力先パス（コピー）
        sheet2_name: Keepaエクスポートのシート名
    """
    source_path = Path(source_path)
    output_path = Path(output_path)

    # ① 元Excelをコピー
    shutil.copy2(source_path, output_path)

    # ② Sheet1からJANコード読み取り
    with open(source_path, "rb") as f:
        df = read_wholesaler_excel(f)

    # ③ EAN → Sheet2行番号のマッピング
    with open(source_path, "rb") as f:
        ean_to_row = build_ean_to_sheet2_row(f, sheet_name=1)

    # ④ Sheet2からタイトル（col1）を取得（数式参照できない固定値用）
    with open(source_path, "rb") as f:
        sheet2_raw = pd.read_excel(
            f, sheet_name=1, header=0, dtype=str, engine="openpyxl",
            usecols=[0, 1, 2],
        )

    row_to_title: dict[int, str] = {}
    for idx, row in sheet2_raw.iterrows():
        excel_row = idx + 2
        title = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        row_to_title[excel_row] = title

    # ⑤ コピーしたExcelのSheet3を書き換え
    wb = openpyxl.load_workbook(output_path)

    if "Sheet3" in wb.sheetnames:
        ws = wb["Sheet3"]
    else:
        ws = wb.create_sheet("Sheet3")

    # Row3以降のデータをクリア（Row1-2のヘッダーは残す）
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # ⑥ JANコードごとにSheet3の行を生成
    sq = f"'{sheet2_name}'"  # シート名参照用（数式内）
    out_row = 3
    unique_jans = df["jan_code"].unique()

    # JANコード → 卸データのマッピング
    jan_to_ws = {}
    for _, row in df.iterrows():
        jan = row["jan_code"]
        if jan not in jan_to_ws:
            jan_to_ws[jan] = row.to_dict()

    matched = 0
    unmatched = 0
    for jan in unique_jans:
        n = out_row
        ws_data = jan_to_ws.get(jan, {})

        if jan in ean_to_row:
            # マッチあり: Sheet2への参照数式
            s2r = ean_to_row[jan]
            matched += 1

            ws.cell(row=n, column=4, value=f"={sq}!AR{s2r}")     # D: 商品名日本語
            title = row_to_title.get(s2r, "")
            ws.cell(row=n, column=5, value=title)                  # E: タイトル
            ws.cell(row=n, column=7, value=f"={sq}!B{s2r}")       # G: ASIN
            ws.cell(row=n, column=9, value=f"={sq}!E{s2r}")       # I: 型番
            ws.cell(row=n, column=11, value=1)                     # K: セット内容
            ws.cell(row=n, column=14, value=f"={sq}!AM{s2r}")     # N: 卸
            ws.cell(row=n, column=15, value=f"={sq}!F{s2r}")      # O: 30日ランク下落
            ws.cell(row=n, column=16, value=f"={sq}!Y{s2r}")      # P: FBAセラー数
            ws.cell(row=n, column=19, value=5)                     # S: 初回仕入れ
            ws.cell(row=n, column=20, value=f"={sq}!O{s2r}")      # T: Buy Box価格
            ws.cell(row=n, column=22, value=f"={sq}!AQ{s2r}")     # V: FBA手数料
            ws.cell(row=n, column=24, value=f"={sq}!AD{s2r}")     # X: Weight

            ws.cell(row=n, column=23, value=f"=K{n}*N{n}*1.1")
            ws.cell(row=n, column=21, value=f"=(W{n}+(X{n}*$AI$2))/$AH$2")
            ws.cell(row=n, column=25, value=f"=(T{n}-U{n}-V{n})")
            ws.cell(row=n, column=26, value=f"=(O{n}/(P{n}+1))*Y{n}")
            ws.cell(row=n, column=27, value=f"=Y{n}/T{n}")
            ws.cell(row=n, column=28, value=f"=Y{n}*S{n}")
            ws.cell(row=n, column=29, value=f"=S{n}*T{n}*$AI$2")
            ws.cell(row=n, column=30, value=f"=V{n}*$AI$2*S{n}")
            ws.cell(row=n, column=31, value=f"=S{n}*W{n}")
            ws.cell(row=n, column=32, value=f"=S{n}*X{n}*$AI$2")
            ws.cell(row=n, column=33, value=f"=AB{n}*$AH$2")
        else:
            # マッチなし: 卸データのみ出力
            unmatched += 1
            ws.cell(row=n, column=4, value=ws_data.get("product_name_jp", ""))
            ws.cell(row=n, column=8, value="Keepaデータなし")
            ws.cell(row=n, column=9, value=ws_data.get("part_number", ""))
            ws.cell(row=n, column=11, value=1)
            ws.cell(row=n, column=14, value=ws_data.get("wholesale_price", 0))
            ws.cell(row=n, column=19, value=5)

        out_row += 1

    wb.save(output_path)
    wb.close()

    return {"total_jans": len(unique_jans), "matched": matched, "output_path": str(output_path)}


def generate_sheet3_from_api(
    wholesaler_df: pd.DataFrame,
    keepa_results: dict[str, dict],
    output_path: str | Path,
    exchange_rate: float = 150.0,
):
    """
    Keepa API/エクスポート結果からExcelを新規作成し、Sheet3形式で書き込む。
    元Sheet3と同一のヘッダー配色・列構成・計算式を再現。
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    output_path = Path(output_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet3"

    # ── ヘッダー配色（元Sheet3と同一） ──
    BLUE = PatternFill("solid", fgColor="0000FF")
    GREEN = PatternFill("solid", fgColor="B6D7A8")
    YELLOW = PatternFill("solid", fgColor="FFF2CC")
    RED = PatternFill("solid", fgColor="FF0000")
    LIME = PatternFill("solid", fgColor="00FF00")
    ORANGE = PatternFill("solid", fgColor="F9CB9C")
    HF = Font(bold=True, size=10)
    NF = Font(size=11)

    # Row2ヘッダー: (列, ヘッダー名, 背景色, 列幅)
    ROW2 = [
        (1,"重複確認",BLUE,13),(2,"日付",BLUE,13),(3,"購入先問屋",BLUE,13),
        (4,"商品名",GREEN,13),(5,"タイトル",GREEN,13),(6,"SKU",GREEN,13),
        (7,"ASIN",GREEN,15.1),(8,"備考",GREEN,13),(9,"販売可能品",GREEN,11.6),
        (10,"購入在庫",GREEN,15.4),(11,"売価最新更新日",GREEN,13),
        (12,"売価90日変化率",GREEN,13),(13,"BBセラー",GREEN,13),
        (14,"出品許可",GREEN,9),(15,"30キーパ",GREEN,13),
        (16,"セラー数（FBA）",GREEN,13),(17,"セラー数（無在庫）",GREEN,13),
        (18,"販売数（自社1M）",YELLOW,13),(19,"初回仕入れ",BLUE,13),
        (20,"売価USD",GREEN,13),(21,"仕入＋送料（FBA）",ORANGE,13),
        (22,"Amazon手数料",GREEN,13),(23,"仕入値",GREEN,9.5),
        (24,"Weight（FBA）",GREEN,13),(25,"損益",RED,13),
        (26,"利益額",RED,13),(27,"利益率",RED,13),
        (28,"利益額2",LIME,13),(29,"売上",LIME,13),(30,"手数料",LIME,13),
        (31,"原価",LIME,10.2),(32,"原価送料",LIME,11.2),(33,"利益/円",LIME,11),
    ]

    # Row1サブヘッダー
    for col, text in [(9,"型番"),(11,"出荷元販売単位"),(12,"セット内容"),
                       (13,"単位"),(14,"1本単価税込"),(15,"1セット料金")]:
        ws.cell(row=1, column=col, value=text).font = NF

    # Row2メインヘッダー
    for col, text, fill, width in ROW2:
        cell = ws.cell(row=2, column=col, value=text)
        cell.font = HF
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    # 定数セル（Row2のAH-AM）
    AH, AI = exchange_rate, 3
    ws.cell(row=2, column=34, value=AH)
    ws.cell(row=2, column=35, value=AI)
    ws.cell(row=2, column=36, value="ポンド")
    ws.cell(row=2, column=37, value=453.59)
    ws.cell(row=2, column=38, value=0.23)
    ws.cell(row=2, column=39, value=106.14006)

    # ── データ行（Row3〜） ──
    from datetime import date
    today = date.today().isoformat()

    unique_jans = wholesaler_df["jan_code"].unique()
    jan_to_ws = {}
    for _, row in wholesaler_df.iterrows():
        if row["jan_code"] not in jan_to_ws:
            jan_to_ws[row["jan_code"]] = row.to_dict()

    matched = 0
    out_row = 3

    for jan in unique_jans:
        kp = keepa_results.get(jan)
        if not kp:
            continue

        # 30日ランク下落がない商品はスキップ
        if not kp.get("sales_rank_drops_30"):
            continue

        ws_data = jan_to_ws.get(jan, {})
        matched += 1
        n = out_row

        N = ws_data.get("wholesale_price", 0)
        T = kp.get("buy_box_price_usd")
        V = kp.get("total_amazon_fee_usd")
        X = kp.get("package_weight_g") or 0
        O = kp.get("sales_rank_drops_30")
        P = kp.get("fba_seller_count") or 0
        S = 5
        L = 1
        W = L * N * 1.1
        U = (W + X * AI) / AH if AH > 0 else 0

        ws.cell(row=n, column=2, value=today)
        ws.cell(row=n, column=4, value=ws_data.get("product_name_jp", ""))
        ws.cell(row=n, column=5, value=kp.get("title", ""))
        ws.cell(row=n, column=7, value=kp.get("asin", ""))
        ws.cell(row=n, column=9, value=ws_data.get("part_number", ""))
        ws.cell(row=n, column=11, value=L)
        ws.cell(row=n, column=13, value=kp.get("buy_box_seller", ""))
        ws.cell(row=n, column=14, value=N)
        ws.cell(row=n, column=15, value=O)
        ws.cell(row=n, column=16, value=P)
        ws.cell(row=n, column=17, value=kp.get("fbm_seller_count"))
        ws.cell(row=n, column=19, value=S)
        ws.cell(row=n, column=20, value=T)
        ws.cell(row=n, column=22, value=V)
        ws.cell(row=n, column=24, value=X if X > 0 else None)

        # 計算値（元Sheet3と同一の計算・丸めなし）
        ws.cell(row=n, column=23, value=W if W else None)
        ws.cell(row=n, column=21, value=U if U else None)

        if T and T > 0 and V is not None:
            Y = T - U - V
            Z = (O / (P + 1)) * Y if O else 0
            AA = Y / T
            AB = Y * S
            AG = AB * AH
            ws.cell(row=n, column=25, value=round(Y, 2))
            ws.cell(row=n, column=26, value=round(Z, 2))
            c = ws.cell(row=n, column=27, value=round(AA, 4))
            c.number_format = "0.00%"
            ws.cell(row=n, column=28, value=round(AB, 2))
            ws.cell(row=n, column=29, value=round(S * T * AI, 2))
            ws.cell(row=n, column=30, value=round(V * AI * S, 2))
            ws.cell(row=n, column=31, value=round(S * W, 2))
            ws.cell(row=n, column=32, value=round(S * X * AI, 2))
            ws.cell(row=n, column=33, value=round(AG, 2))

        out_row += 1

    wb.save(output_path)
    wb.close()
    return {"total_jans": len(unique_jans), "matched": matched, "output_path": str(output_path)}


def build_dashboard_data(
    wholesaler_df: pd.DataFrame,
    ean_to_row: dict[str, int],
    file_buffer,
    sheet_name: str | int = 1,
    exchange_rate: float = 150.0,
) -> pd.DataFrame:
    """
    Sheet1 + Sheet2 のデータを結合し、利益計算済みのDataFrameを返す。
    ダッシュボード表示・フィルタリング用。
    """
    raw = pd.read_excel(
        file_buffer, sheet_name=sheet_name,
        header=0, dtype=str, engine="openpyxl",
    )

    # Sheet2の行番号 → データ辞書
    row_to_data: dict[int, dict] = {}
    for idx, row in raw.iterrows():
        excel_row = idx + 2
        row_to_data[excel_row] = {
            "title": str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else "",
            "asin": str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else "",
            "buy_box_usd": _safe_float(row.iloc[14]),
            "drops_30": _safe_float(row.iloc[5]),
            "fba_sellers": _safe_float(row.iloc[24]) or 0,
            "fbm_sellers": _safe_float(row.iloc[25]) or 0,
            "weight_g": _safe_float(row.iloc[29]) or 0,
            "amazon_fee": _safe_float(row.iloc[42]) if len(row) > 42 else None,
            "wholesale_keepa": _safe_float(row.iloc[38]) if len(row) > 38 else None,
        }

    AH = exchange_rate
    AI = 3
    results = []

    for jan in wholesaler_df["jan_code"].unique():
        if jan not in ean_to_row:
            continue
        s2r = ean_to_row[jan]
        s2 = row_to_data.get(s2r)
        if not s2:
            continue

        ws = wholesaler_df[wholesaler_df["jan_code"] == jan].iloc[0]

        # N値（Keepa卸優先、乖離大なら卸データ）
        keepa_ws = s2.get("wholesale_keepa") or 0
        input_ws = ws.get("wholesale_price", 0)
        if keepa_ws > 0 and input_ws > 0:
            ratio = max(keepa_ws, input_ws) / min(keepa_ws, input_ws)
            N = keepa_ws if ratio < 3 else input_ws
        else:
            N = keepa_ws or input_ws

        T = s2["buy_box_usd"]
        V = s2["amazon_fee"]
        X = s2["weight_g"]
        W = N * 1.1
        U = (W + X * AI) / AH if AH > 0 else 0
        Y = (T - U - V) if T and V else None
        P = s2["fba_sellers"]
        O = s2["drops_30"]
        AA = (Y / T) if Y is not None and T and T > 0 else None

        results.append({
            "JAN": jan,
            "ASIN": s2["asin"],
            "タイトル": s2["title"][:60],
            "商品名": ws.get("product_name_jp", "")[:30],
            "売価USD": T,
            "仕入値(税込)": round(W) if W else None,
            "仕入+送料USD": round(U, 2) if U else None,
            "手数料USD": V,
            "損益USD": round(Y, 2) if Y is not None else None,
            "利益率": round(AA, 4) if AA is not None else None,
            "30日drop": O,
            "FBAセラー": int(P),
            "FBMセラー": int(s2["fbm_sellers"]),
            "重量g": int(X) if X else None,
        })

    return pd.DataFrame(results)


def _safe_float(val) -> float | None:
    try:
        v = float(val)
        return v if v > 0 else None
    except (TypeError, ValueError):
        return None
