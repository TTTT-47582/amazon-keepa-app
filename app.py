"""
JAN コード → US Amazon リサーチツール

卸問屋のJANコードExcelをアップロードすると、
Sheet2のKeepaデータとマッチングし、
Sheet3に数式ごと自動生成する。

起動コマンド: streamlit run app.py
"""

from __future__ import annotations

import os
import tempfile
import subprocess
import platform
from pathlib import Path
import requests
import streamlit as st
from src.excel_io import read_wholesaler_excel, build_ean_to_sheet2_row, generate_sheet3, generate_sheet3_from_api, read_keepa_export_as_results
from src.keepa_client import query_jan_codes_us, estimate_tokens, _get_api, _resolve_api_key

st.set_page_config(
    page_title="Amapro - 仕入れリサーチ",
    page_icon="static/favicon-32.png",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Amazon風カラーテーマ
st.markdown("""
<style>
    /* ヘッダーバー */
    header[data-testid="stHeader"] {
        background-color: #131921 !important;
    }
    /* 不要なヘッダーアイコン（星・ペン・GitHub）を非表示 */
    header [data-testid="stToolbar"] {
        display: none !important;
    }
    /* サイドバーを常に表示 */
    section[data-testid="stSidebar"] {
        display: block !important;
        width: 300px !important;
        min-width: 300px !important;
        transform: none !important;
        position: relative !important;
    }
    button[data-testid="stSidebarCollapseButton"] {
        display: none !important;
    }
    [data-testid="collapsedControl"] {
        display: none !important;
    }
    /* サイドバー */
    section[data-testid="stSidebar"] {
        background-color: #232F3E !important;
    }
    section[data-testid="stSidebar"] * {
        color: #FFFFFF !important;
    }
    section[data-testid="stSidebar"] .stSelectbox label,
    section[data-testid="stSidebar"] .stNumberInput label,
    section[data-testid="stSidebar"] .stTextInput label {
        color: #FF9900 !important;
    }
    section[data-testid="stSidebar"] input,
    section[data-testid="stSidebar"] .stNumberInput input,
    section[data-testid="stSidebar"] .stTextInput input {
        color: #0F1111 !important;
        background-color: #FFFFFF !important;
    }
    section[data-testid="stSidebar"] hr {
        border-color: #3B4859 !important;
    }
    /* デフォルトのページナビを非表示（カスタムナビに置き換え） */
    [data-testid="stSidebarNav"] {
        display: none !important;
    }
    /* メインエリア */
    .main .block-container {
        background-color: #FFFFFF;
    }
    /* タイトル */
    h1 {
        color: #131921 !important;
    }
    /* プライマリボタン → Amazon オレンジ */
    button[kind="primary"], .stButton > button[kind="primary"] {
        background-color: #FF9900 !important;
        border-color: #E88B00 !important;
        color: #131921 !important;
        font-weight: bold !important;
    }
    button[kind="primary"]:hover {
        background-color: #E88B00 !important;
    }
    /* ダウンロードボタン */
    .stDownloadButton > button {
        background-color: #FFD814 !important;
        border-color: #FCD200 !important;
        color: #0F1111 !important;
        font-weight: bold !important;
    }
    .stDownloadButton > button:hover {
        background-color: #F7CA00 !important;
    }
    /* 成功メッセージ */
    .stSuccess {
        background-color: #F0FFF0 !important;
        border-left-color: #FF9900 !important;
    }
    /* メトリクス */
    [data-testid="stMetricValue"] {
        color: #131921 !important;
        font-weight: bold !important;
    }
    [data-testid="stMetricLabel"] {
        color: #565959 !important;
    }
    /* ファイルアップローダー */
    [data-testid="stFileUploader"] {
        border-color: #FF9900 !important;
    }
    /* リンク色 */
    a {
        color: #007185 !important;
    }
    a:hover {
        color: #C7511F !important;
    }
</style>
""", unsafe_allow_html=True)


@st.cache_data(ttl=3600)
def _fetch_exchange_rate() -> tuple[float, str]:
    """Frankfurter API から USD→JPY レートを取得する"""
    try:
        resp = requests.get(
            "https://api.frankfurter.app/latest?from=USD&to=JPY",
            timeout=5,
        )
        resp.raise_for_status()
        data = resp.json()
        return float(data["rates"]["JPY"]), data.get("date", "")
    except Exception:
        return 150.0, ""


def main():
    # メインロゴ（ライト背景用）
    import base64
    with open("static/logo-lockup.png", "rb") as f:
        logo_b64 = base64.b64encode(f.read()).decode()
    st.markdown(
        f'<div style="text-align:center; padding:10px 0 40px 0;">'
        f'<img src="data:image/png;base64,{logo_b64}" height="90" alt="amapro">'
        f'</div>',
        unsafe_allow_html=True,
    )

    # ── サイドバー ──
    with st.sidebar:
        # サイドバーロゴ + ナビ
        with open("static/logo-lockup-dark.png", "rb") as f:
            sidebar_logo_b64 = base64.b64encode(f.read()).decode()
        st.markdown(f"""
        <div style="background:#37475A; margin:-1rem -1rem 1rem -1rem; padding:16px;">
            <div style="text-align:center; margin-bottom:12px;">
                <a href="/" target="_self"><img src="data:image/png;base64,{sidebar_logo_b64}" height="40" alt="amapro" style="cursor:pointer;"></a>
            </div>
            <a href="/使い方" target="_self"
               style="color:#DDDDDD; text-decoration:none; font-size:13px;
                      display:block; padding:6px 0 6px 15px;">
                📖 使い方ガイド
            </a>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<h3 style='color:#FF9900;'>⚙️ 設定</h3>", unsafe_allow_html=True)

        data_mode = st.radio(
            "データ取得モード",
            ["📄 Excel内Keepaデータ（Sheet2）", "🌐 Keepa APIで取得"],
            index=0,
        )
        use_api = data_mode.startswith("🌐")

        api_key = ""
        max_items = 100
        if use_api:
            env_key = os.getenv("KEEPA_API_KEY", "")
            api_key = st.text_input("Keepa API キー", value=env_key, type="password")
            max_items = st.number_input("処理上限件数", min_value=10, max_value=5000, value=100, step=50)

        st.divider()

        live_rate, rate_date = _fetch_exchange_rate()
        if rate_date:
            st.caption(f"💱 現在レート: 1 USD = ¥{live_rate:.2f}（{rate_date}）")

    # ── メインエリア ──
    if use_api:
        st.markdown("""
        <div style="background:#F7F8FA; border:1px solid #D5D9D9; border-radius:8px;
                    padding:20px 24px; margin-bottom:16px;">
            <h4 style="color:#0F1111; margin:0 0 4px 0;">JANコードExcelをアップロード</h4>
            <p style="color:#565959; font-size:13px; margin:0;">Keepa APIで自動取得します</p>
        </div>
        """, unsafe_allow_html=True)

        uploaded = st.file_uploader("JANコードExcel", type=["xlsx"], label_visibility="collapsed")
        uploaded_keepa = None
    else:
        st.markdown("""
        <div style="background:#F7F8FA; border:1px solid #D5D9D9; border-radius:8px;
                    padding:20px 24px; margin-bottom:16px;">
            <h4 style="color:#0F1111; margin:0 0 4px 0;">Excelファイルをアップロード</h4>
            <p style="color:#565959; font-size:13px; margin:0;">
                Keepaエクスポートは複数ファイルOK（10分割でもまとめて処理）
            </p>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("**① 卸データExcel**")
        uploaded = st.file_uploader("卸データExcel", type=["xlsx"], key="jan_file", label_visibility="collapsed")
        st.markdown("**② Keepaエクスポート**（複数ファイル可）")
        uploaded_keepa_files = st.file_uploader(
            "Keepaエクスポート", type=["xlsx"], key="keepa_files",
            label_visibility="collapsed", accept_multiple_files=True,
        )

    # 後方互換: uploaded_keepa_files が未定義の場合（APIモード）
    if not use_api:
        uploaded_keepa = uploaded_keepa_files if uploaded_keepa_files else None
    else:
        uploaded_keepa = None

    if uploaded is None and not uploaded_keepa:
        return
    if uploaded is None:
        # Keepaファイルだけの場合、最初のファイルをメインとして扱う
        uploaded = uploaded_keepa[0] if isinstance(uploaded_keepa, list) and uploaded_keepa else None
        if uploaded is None:
            return

    # Excel読み込み
    keepa_names = ""
    if isinstance(uploaded_keepa, list):
        keepa_names = "_".join(f.name for f in uploaded_keepa)
    elif uploaded_keepa:
        keepa_names = uploaded_keepa.name
    cache_key = f"{uploaded.name}_{keepa_names or 'none'}"

    if "uploaded_df" not in st.session_state or st.session_state.get("uploaded_cache_key") != cache_key:
        with st.spinner("Excel を読み込み中..."):
            try:
                df = read_wholesaler_excel(uploaded)

                ean_to_row = {}
                keepa_sheet_idx = None
                if not use_api:
                    keepa_file_list = []
                    if isinstance(uploaded_keepa, list) and uploaded_keepa:
                        keepa_file_list = uploaded_keepa
                    elif uploaded_keepa:
                        keepa_file_list = [uploaded_keepa]

                    if keepa_file_list:
                        # 複数Keepaファイルを結合
                        for kf in keepa_file_list:
                            kf.seek(0)
                            partial = build_ean_to_sheet2_row(kf, sheet_name=0)
                            ean_to_row.update(partial)
                        keepa_sheet_idx = "separate"
                        st.caption(f"📁 Keepaファイル {len(keepa_file_list)} 個を結合")
                    else:
                        uploaded.seek(0)
                        import openpyxl as _xl
                        _wb = _xl.load_workbook(uploaded, read_only=True)
                        sheet_count = len(_wb.sheetnames)
                        _wb.close()

                        if sheet_count >= 2:
                            uploaded.seek(0)
                            ean_to_row = build_ean_to_sheet2_row(uploaded, sheet_name=1)
                            keepa_sheet_idx = 1
                        else:
                            uploaded.seek(0)
                            ean_to_row = build_ean_to_sheet2_row(uploaded, sheet_name=0)
                            keepa_sheet_idx = 0

                    # JANコードが読めなかった場合、EANをJANとして使う
                    if df.empty and ean_to_row:
                        import pandas as _pd
                        df = _pd.DataFrame({
                            "jan_code": list(ean_to_row.keys()),
                            "product_name_jp": "",
                            "part_number": "",
                            "retail_price": 0,
                            "wholesale_price": 0,
                            "box_qty": 1,
                        })
                    elif df.empty:
                        st.error("JANコードが見つかりませんでした。")
                        return
            except Exception as e:
                st.error(f"Excel読み込みエラー: {e}")
                return
            st.session_state["uploaded_df"] = df
            st.session_state["ean_to_row"] = ean_to_row
            st.session_state["uploaded_cache_key"] = cache_key
            st.session_state["uploaded_keepa"] = uploaded_keepa
            st.session_state["keepa_sheet_idx"] = keepa_sheet_idx
            st.session_state["generated"] = False

    df = st.session_state.get("uploaded_df")
    if df is None:
        return

    unique_jans = df["jan_code"].nunique()
    st.success(f"✅ **{len(df):,} 行**読み込み（ユニーク JAN: **{unique_jans:,} 件**）")

    if not use_api:
        ean_to_row = st.session_state.get("ean_to_row") or {}
        matched = sum(1 for jan in df["jan_code"].unique() if jan in ean_to_row)
        st.info(f"🔍 Keepaデータとマッチ: **{matched:,} 件** / {unique_jans:,} 件（未マッチも全件出力されます）")

    with st.expander("📋 データプレビュー（先頭20行）", expanded=False):
        st.dataframe(df.head(20), use_container_width=True)

    # APIモード: トークン情報表示
    if use_api:
        process_count = min(unique_jans, max_items)
        tokens = estimate_tokens(process_count)
        st.info(
            f"🔍 処理対象: **{process_count:,} 件**\n\n"
            f"💰 推定トークン: **{tokens:,}**　｜　⏱ 約{max(tokens // 50, 1)}分"
        )

    # 処理実行
    btn_label = "🚀 Sheet3 を自動生成" if not use_api else "🚀 Keepa API で取得＆生成"
    start_btn = st.button(btn_label, type="primary", use_container_width=True)

    if start_btn:
        if use_api:
            if not api_key:
                st.error("Keepa API キーを入力してください")
                return
            try:
                api = _get_api(_resolve_api_key(api_key))
                tokens_left = api.tokens_left
                needed = estimate_tokens(min(unique_jans, max_items))
                st.write(f"🔑 トークン残高: **{tokens_left}** / 必要: **{needed}**")
                if tokens_left < needed:
                    wait_min = max((needed - tokens_left) // 50, 1)
                    st.warning(
                        f"トークンが不足していますが、**自動で待機しながら処理**します。\n\n"
                        f"不足分: {needed - tokens_left} トークン → 約{wait_min}分の追加待ち時間"
                    )
            except Exception as e:
                st.warning(f"トークン確認失敗: {e}")
            _run_api_generation(df, api_key, max_items, live_rate)
        else:
            keepa_idx = st.session_state.get("keepa_sheet_idx")
            if keepa_idx == "separate":
                _run_keepa_multi_generation(uploaded_keepa, df, live_rate)
            elif keepa_idx == 0:
                _run_keepa_multi_generation([uploaded], df, live_rate)
            elif keepa_idx == 1:
                _run_generation(uploaded, None)

    if st.session_state.get("generated"):
        output_path = st.session_state.get("output_path")
        result = st.session_state.get("gen_result", {})
        _show_result(output_path, result)

        # スクリーニング機能
        keepa_data = st.session_state.get("keepa_results_data")
        if keepa_data:
            _show_screening(keepa_data, df, live_rate)


def _run_keepa_multi_generation(keepa_files, df, exchange_rate: float):
    """複数のKeepaエクスポートファイルを結合してSheet3を生成"""
    if isinstance(keepa_files, list):
        file_list = keepa_files
    else:
        file_list = [keepa_files]

    keepa_results = {}
    with st.spinner(f"Keepaデータを読み込み中... ({len(file_list)}ファイル)"):
        for i, kf in enumerate(file_list):
            kf.seek(0)
            partial = read_keepa_export_as_results(kf, sheet_name=0)
            keepa_results.update(partial)
            if len(file_list) > 1:
                st.caption(f"  ファイル {i+1}/{len(file_list)}: {len(partial):,}件読み込み")

    if not keepa_results:
        st.error("Keepaデータからマッチする商品が見つかりませんでした。")
        return

    st.info(f"📊 Keepa合計: **{len(keepa_results):,}件** のEANを読み込み")

    desktop = Path.home() / "Desktop"
    if desktop.exists():
        output_path = desktop / "リサーチ結果.xlsx"
    else:
        output_path = Path(tempfile.gettempdir()) / "リサーチ結果.xlsx"

    with st.spinner("Sheet3 を生成中..."):
        try:
            result = generate_sheet3_from_api(df, keepa_results, output_path, exchange_rate)
        except Exception as e:
            st.error(f"生成エラー: {e}")
            return

    found = sum(1 for r in keepa_results.values() if r.get("found"))
    st.session_state["generated"] = True
    st.session_state["output_path"] = str(output_path)
    st.session_state["gen_result"] = result
    st.session_state["keepa_results_data"] = keepa_results

    st.success(
        f"🎉 完了！ **{found:,} 件**のUS Amazon商品データを出力 "
        f"（全{result['total_jans']:,} JAN中）"
    )


def _run_generation(uploaded_file, keepa_file=None):
    """元Excelをコピーし、Sheet3を自動生成する。Keepaが別ファイルなら結合する。"""
    with st.spinner("Sheet3 を生成中..."):
        uploaded_file.seek(0)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_in:
            tmp_in.write(uploaded_file.read())
            tmp_in_path = tmp_in.name

        # Keepaが別ファイルの場合、Sheet2として結合
        if keepa_file:
            import openpyxl
            keepa_file.seek(0)
            keepa_wb = openpyxl.load_workbook(keepa_file)
            keepa_ws = keepa_wb.active
            keepa_sheet_name = keepa_ws.title

            main_wb = openpyxl.load_workbook(tmp_in_path)
            new_ws = main_wb.create_sheet(keepa_sheet_name)
            for row in keepa_ws.iter_rows(values_only=True):
                new_ws.append(list(row))
            main_wb.save(tmp_in_path)
            main_wb.close()
            keepa_wb.close()

        desktop = Path.home() / "Desktop"
        if desktop.exists():
            output_path = desktop / "リサーチ結果.xlsx"
        else:
            output_path = Path(tempfile.gettempdir()) / "リサーチ結果.xlsx"

        import openpyxl
        wb_tmp = openpyxl.load_workbook(tmp_in_path, read_only=True)
        sheet2_name = wb_tmp.sheetnames[1] if len(wb_tmp.sheetnames) > 1 else "Sheet2"
        wb_tmp.close()

        try:
            result = generate_sheet3(
                source_path=tmp_in_path,
                output_path=output_path,
                sheet2_name=sheet2_name,
            )
        except Exception as e:
            st.error(f"生成エラー: {e}")
            import traceback
            st.code(traceback.format_exc())
            return
        finally:
            os.unlink(tmp_in_path)

    st.session_state["generated"] = True
    st.session_state["output_path"] = str(output_path)
    st.session_state["gen_result"] = result

    st.success(
        f"🎉 完了！ **{result['matched']:,} 件**をSheet3に生成 "
        f"（全{result['total_jans']:,} JAN中）"
    )


def _run_api_generation(df, api_key: str, max_items: int, exchange_rate: float):
    """Keepa APIでJANコードを取得してSheet3を生成する（トークン不足時は自動待機）"""
    unique_jans = df["jan_code"].unique().tolist()[:max_items]

    progress_bar = st.progress(0, text="Keepa API に問い合わせ中...")
    status_text = st.empty()

    def on_progress(done, total):
        progress_bar.progress(done / total, text=f"処理中... {done:,}/{total:,} 件（トークン不足時は自動待機）")
        status_text.caption(f"完了: {done:,} / {total:,}")

    try:
        keepa_results = query_jan_codes_us(
            unique_jans, api_key=api_key,
            batch_size=50,  # 小刻みにしてトークン補充の猶予を確保
            use_offers=False,
            progress_callback=on_progress,
        )
    except Exception as e:
        st.error(f"Keepa API エラー: {e}")
        return

    progress_bar.progress(1.0, text="完了！")

    desktop = Path.home() / "Desktop"
    if desktop.exists():
        output_path = desktop / "リサーチ結果.xlsx"
    else:
        output_path = Path(tempfile.gettempdir()) / "リサーチ結果.xlsx"

    try:
        result = generate_sheet3_from_api(df, keepa_results, output_path, exchange_rate)
    except Exception as e:
        st.error(f"Excel生成エラー: {e}")
        return

    found = sum(1 for r in keepa_results.values() if r.get("found"))
    st.session_state["generated"] = True
    st.session_state["output_path"] = str(output_path)
    st.session_state["gen_result"] = result

    st.success(
        f"🎉 完了！ **{found:,} 件**がUS Amazonに出品あり "
        f"（全{len(unique_jans):,} JAN中）"
    )


def _show_result(output_path: str, result: dict):
    """結果表示とファイルオープン"""
    st.divider()

    c1, c2, c3 = st.columns(3)
    c1.metric("全JANコード", f"{result.get('total_jans', 0):,}")
    c2.metric("Sheet2マッチ", f"{result.get('matched', 0):,}")
    c3.metric("未マッチ", f"{result.get('total_jans', 0) - result.get('matched', 0):,}")

    st.success(f"📂 保存先: **{output_path}**")

    # 自動でExcelを開く
    try:
        if platform.system() == "Darwin":
            subprocess.Popen(["open", output_path])
        elif platform.system() == "Windows":
            subprocess.Popen(["start", "", output_path], shell=True)
    except Exception:
        pass

    # ダウンロードボタンも用意
    try:
        with open(output_path, "rb") as f:
            st.download_button(
                "📥 結果 Excel を手動ダウンロード",
                data=f.read(),
                file_name="リサーチ結果.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    except Exception:
        pass


def _show_screening(keepa_data: dict, df, exchange_rate: float):
    """スクリーニング機能: スライダーで条件を指定して絞り込み"""
    import pandas as pd

    st.divider()
    st.markdown("""
    <h2 style="color:#131921;">🔍 スクリーニング</h2>
    <p style="color:#565959;">条件を指定して商品を絞り込み、スクリーニング済みExcelをダウンロードできます</p>
    """, unsafe_allow_html=True)

    jan_to_ws = {}
    for _, row in df.iterrows():
        if row["jan_code"] not in jan_to_ws:
            jan_to_ws[row["jan_code"]] = row.to_dict()

    AH, AI = exchange_rate, 3
    items = []
    for jan, kp in keepa_data.items():
        if not kp:
            continue
        ws = jan_to_ws.get(jan, {})
        N = ws.get("wholesale_price", 0)
        T = kp.get("buy_box_price_usd")
        V = kp.get("total_amazon_fee_usd")
        X = kp.get("package_weight_g") or 0
        O = kp.get("sales_rank_drops_30")
        P = kp.get("fba_seller_count") or 0
        W = N * 1.1
        U = (W + X * AI) / AH if AH > 0 else 0
        Y = (T - U - V) if T and V else None
        AA = (Y / T * 100) if Y is not None and T and T > 0 else None

        items.append({
            "JAN": jan,
            "ASIN": kp.get("asin", ""),
            "タイトル": kp.get("title", "")[:50],
            "商品名": ws.get("product_name_jp", "")[:30],
            "売価USD": T,
            "仕入値": round(W) if W else 0,
            "損益USD": round(Y, 2) if Y else None,
            "利益率%": round(AA, 1) if AA else None,
            "30日drop": O,
            "FBAセラー": P,
            "FBMセラー": kp.get("fbm_seller_count") or 0,
            "重量g": X,
            "Amazon販売": kp.get("is_amazon_selling", False),
            "Amazon価格": kp.get("amazon_price_usd"),
            "AmazonBB%": kp.get("amazon_bb_pct_90"),
        })

    all_df = pd.DataFrame(items)
    if all_df.empty:
        return

    # ── フィルタUI ──
    st.markdown("""
    <div style="background:#F7F8FA; border:1px solid #D5D9D9; border-radius:8px;
                padding:16px 20px; margin-bottom:16px;">
        <h4 style="color:#FF9900; margin:0 0 8px 0;">フィルタ条件</h4>
    </div>
    """, unsafe_allow_html=True)

    drops_min = st.slider(
        "30日ランク下落（最低回数）",
        min_value=0, max_value=200, value=19, step=1,
        help="数値が大きいほど売れている商品",
    )
    margin_min = st.slider(
        "利益率（最低 %）",
        min_value=-50, max_value=100, value=10, step=5,
        help="損益 ÷ 売価 × 100",
    )
    exclude_amazon = st.checkbox(
        "🚫 Amazon本体が販売している商品を除外",
        value=True,
        help="Amazon.comが直接販売・Buy Boxを持っている商品を除外",
    )

    # ── フィルタ適用 ──
    filtered = all_df.copy()
    filtered = filtered[filtered["売価USD"].notna() & (filtered["売価USD"] > 0)]

    if drops_min > 0:
        filtered = filtered[filtered["30日drop"].notna() & (filtered["30日drop"] >= drops_min)]
    if margin_min != -50:
        filtered = filtered[filtered["利益率%"].notna() & (filtered["利益率%"] >= margin_min)]
    if exclude_amazon:
        filtered = filtered[~filtered["Amazon販売"]]

    # ── 結果表示 ──
    st.divider()
    profitable = filtered[filtered["損益USD"].notna() & (filtered["損益USD"] > 0)]

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("全商品", f"{len(all_df[all_df['売価USD'].notna()]):,}")
    m2.metric("スクリーニング後", f"{len(filtered):,}")
    m3.metric("利益あり", f"{len(profitable):,}")
    if not profitable.empty:
        m4.metric("平均利益率", f"{profitable['利益率%'].mean():.1f}%")
    else:
        m4.metric("平均利益率", "-")

    sort_col = st.selectbox(
        "並び替え",
        ["利益率%（高い順）", "損益USD（高い順）", "30日drop（多い順）", "FBAセラー（少ない順）", "売価USD（高い順）"],
    )
    sort_map = {
        "利益率%（高い順）": ("利益率%", False),
        "損益USD（高い順）": ("損益USD", False),
        "30日drop（多い順）": ("30日drop", False),
        "FBAセラー（少ない順）": ("FBAセラー", True),
        "売価USD（高い順）": ("売価USD", False),
    }
    sort_key, ascending = sort_map[sort_col]
    filtered = filtered.sort_values(sort_key, ascending=ascending, na_position="last")

    # ASINにAmazonリンクを付与
    filtered = filtered.copy()
    filtered["Amazon"] = filtered["ASIN"].apply(
        lambda x: f"https://www.amazon.com/dp/{x}" if x else ""
    )
    display_cols = ["ASIN", "タイトル", "売価USD", "仕入値", "損益USD", "利益率%",
                    "FBAセラー", "30日drop", "Amazon"]
    st.dataframe(
        filtered[display_cols],
        use_container_width=True,
        height=500,
        column_config={
            "Amazon": st.column_config.LinkColumn("Amazon", display_text="開く"),
        },
    )

    # スクリーニング済みExcelダウンロード
    if not filtered.empty:
        from src.excel_io import generate_sheet3_from_api
        filtered_keepa = {row["JAN"]: keepa_data[row["JAN"]] for _, row in filtered.iterrows() if row["JAN"] in keepa_data}
        buf_path = Path(tempfile.gettempdir()) / "スクリーニング結果.xlsx"
        generate_sheet3_from_api(df, filtered_keepa, buf_path, exchange_rate)
        with open(buf_path, "rb") as f:
            st.download_button(
                f"📥 スクリーニング済み Excel（{len(filtered):,}件）",
                data=f.read(),
                file_name="スクリーニング結果.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )


if __name__ == "__main__":
    main()
